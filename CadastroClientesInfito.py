from openpyxl import Workbook, load_workbook
import os

arquivo = "tabelaclientes.xlsx"  # Define o nome do arquivo Excel que será usado

if os.path.exists(arquivo):  # Verifica se o arquivo já existe na pasta
    wb = load_workbook(arquivo)  # Abre o arquivo Excel existente
    ws = wb.active  # Seleciona a aba ativa (planilha principal)
else:
    wb = Workbook()  # Cria um novo arquivo Excel (novo "workbook")
    ws = wb.active  # Seleciona a aba padrão criada automaticamente
    ws.append(["Nome", "Email"])  # Adiciona a primeira linha como cabeçalho
while True:
    while True:
        opcao = input(
"""\n1 - Cadastrar clientes.
2 - Lista de clientes.
3 - Buscar clientes.
4 - Sair.
Escolha uma das opções: """)
        if opcao in ['1', '2', '3', '4']:
            break
        print("OPÇÃO INVÁLIDA.")
        print("_______________________________")

    #CADASTRAR CLIENTES
    if opcao == '1':
        print("_______________________________")
        print("CADASTRO DE CLIENTE")

        while True:

            cancelado = False

            while True:
                nome = input("Digite o nome do cliente: ").title()

                if nome.lower() == "cancelar":
                    cancelado = True
                    break

                encontrou = False

                for nome_ws, contato in list(ws.values)[1:]:
                    if nome == nome_ws:
                        print("Não é possível cadastrar esse nome. Motivo: o nome já está cadastrado!")
                        encontrou = True
                        break
                if encontrou == False:
                    break

            if cancelado:
                print("Cadastro cancelado.")
                break

            while True:
                contato = input("Digite o e-mail do cliente: ")

                if contato == "cancelar":
                    cancelado = True
                    break

                if "@" in contato and "." in contato:
                    break
                print("Email inválido")
            if cancelado:
                print("Cadastro cancelado")
                break
            ws.append([nome, contato])

            while True:
                continuar = input("Deseja cadastrar outro cliente? (s/n): ").lower()
                if continuar in ["s", "n"]:
                    break
                print("Resposta inválida.")

            if continuar == "n":
                break

            wb.save(arquivo)
        print("\nClientes salvos com sucesso!")
        print("_______________________________")

    #LISTA DE CLIENTES
    elif opcao == '2':
        print("_______________________________")
        print("\nLISTA DE CLIENTES\n")

        for nome, contato in list(ws.values)[1:]:
            print(f"{nome} / {contato}")

        print("_______________________________")

    #PESQUISAR CLIENTE
    elif opcao == '3':
        print("____________________________")
        print("\nBUSCAR CLIENTES\n")

        busca = input("Digite um nome que deseja pesquisar: ").title()

        encontrou = False

        for nome, contato in list(ws.values)[1:]:
            if nome == busca:
                print("Encontrado!")
                encontrou = True
        if encontrou == False:
            print("Nome não encontrado.")
        print("____________________________")

    #FINALIZAR
    elif opcao == '4':
        wb.save(arquivo)
        print("_______________________________")
        print("Arquivo salvo!\nFechando...")
        break