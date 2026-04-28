from openpyxl import Workbook, load_workbook
import os

from clientes import cadastrar_clientes, listar_clientes, buscar_cliente, fechar, deletar_cliente

arquivo = "data/tabelaclientes.xlsx"

if os.path.exists(arquivo):
    wb = load_workbook(arquivo)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["Nome", "Email"])

while True:
    while True:
        opcao = input(
"""MENU
1 - Cadastrar clientes.
2 - Lista de clientes.
3 - Buscar clientes.
4 - Sair.
5 - Deletar cliente.
Escolha uma das opções: """
        )
        if opcao in ['1', '2', '3', '4', '5']:
            break
        print("OPÇÃO INVÁLIDA.")
        print("_______________________________")

    # CADASTRAR CLIENTES
    if opcao == '1':
        cadastrar_clientes(ws, wb, arquivo)


    # LISTA DE CLIENTES
    elif opcao == '2':
        listar_clientes(ws)

    # PESQUISAR CLIENTE
    elif opcao == '3':
        buscar_cliente(ws)

    # FINALIZAR
    elif opcao == '4':
        fechar(wb, arquivo)
        break

    elif opcao == '5':
        deletar_cliente(ws, wb, arquivo)
